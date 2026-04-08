from io import BytesIO
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from django.http import HttpResponse

from apps.dashboards import selectors


def _format_decimal(value):
    """Formata valores numericos para duas casas decimais."""
    if value is None:
        return 0.0
    return float(value)


def _format_percentage(value):
    """Formata percentagens com duas casas decimais."""
    if value is None:
        return '0.00%'
    return f'{float(value):.2f}%'


def _build_filename(day: date) -> str:
    """Constroi nome de ficheiro consistente para export dia anterior."""
    return f'dia_anterior_{day:%Y%m%d}.xlsx'


def _setup_worksheet_styles(worksheet):
    """Aplica estilos básicos ao worksheet."""
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    return header_fill, header_font, border


def export_previous_day_excel(payload: dict, filters: dict = None) -> HttpResponse:
    """
    Exporta relatório do Dia anterior para Excel com múltiplas sheets.

    Sheets:
    1. Resumo - KPIs principais
    2. Assistentes - Top e bottom assistentes
    3. Tipificações - Breakdown de tipificações
    4. Ações - Resumo de ações
    5. Auditoria - Chamadas priorizadas
    6. Insights - Insights automáticos com ações sugeridas
    """
    filename = _build_filename(payload['day'])

    workbook = Workbook()
    workbook.remove(workbook.active)  # Remove sheet default

    # Sheet 1 - Resumo
    _build_resumo_sheet(workbook, payload)

    # Sheet 2 - Assistentes
    _build_assistentes_sheet(workbook, payload)

    # Sheet 3 - Tipificações
    _build_tipificacoes_sheet(workbook, payload)

    # Sheet 4 - Ações
    _build_acoes_sheet(workbook, payload, filters)

    # Sheet 5 - Auditoria
    _build_auditoria_sheet(workbook, payload)

    # Sheet 6 - Insights
    _build_insights_sheet(workbook, payload)

    # Autofit columns (aproximação)
    for sheet in workbook.sheetnames:
        ws = workbook[sheet]
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _build_resumo_sheet(workbook: Workbook, payload: dict):
    """Sheet 1: Resumo com KPIs principais."""
    ws = workbook.create_sheet('Resumo')

    headers = ['Metrica', 'Valor']
    ws.append(headers)
    _setup_worksheet_styles(ws)

    kpis = payload['kpis']
    rows = [
        ['Data', payload['day'].strftime('%Y-%m-%d')],
        ['Total de chamadas', int(kpis['total_calls'])],
        ['Taxa de retencao (%)', _format_decimal(kpis['retention_rate'])],
        ['Taxa nao retencao (%)', _format_decimal(kpis['non_retention_rate'])],
        ['Percentagem sem acao (%)', _format_decimal(kpis['no_action_pct'])],
        ['Taxa de inconsistencias (%)', _format_decimal(kpis['inconsistency_rate'])],
    ]

    for row in rows:
        ws.append(row)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal='left')


def _build_assistentes_sheet(workbook: Workbook, payload: dict):
    """Sheet 2: Ranking de assistentes (top e bottom)."""
    ws = workbook.create_sheet('Assistentes')

    headers = ['Nome', 'Total chamadas', 'Taxa retencao (%)', 'Ranking']
    ws.append(headers)
    _setup_worksheet_styles(ws)

    rows = []

    for assistant in payload['assistants']['top']:
        rows.append([
            assistant['assistant_name'],
            int(assistant['total_calls']),
            _format_decimal(assistant['retention_rate']),
            'Top',
        ])

    for assistant in payload['assistants']['bottom']:
        rows.append([
            assistant['assistant_name'],
            int(assistant['total_calls']),
            _format_decimal(assistant['retention_rate']),
            'Bottom',
        ])

    for row in rows:
        ws.append(row)


def _build_tipificacoes_sheet(workbook: Workbook, payload: dict):
    """Sheet 3: Breakdown de tipificações."""
    ws = workbook.create_sheet('Tipificacoes')

    headers = ['Tipificacao', 'Total casos', 'Taxa retencao (%)']
    ws.append(headers)
    _setup_worksheet_styles(ws)

    # Adiciona best e worst se existirem
    if payload['tipification']['best']:
        ws.append([
            payload['tipification']['best']['tipification_label'],
            int(payload['tipification']['best']['total_calls']),
            _format_decimal(payload['tipification']['best']['retention_rate']),
        ])

    if payload['tipification']['worst']:
        ws.append([
            payload['tipification']['worst']['tipification_label'],
            int(payload['tipification']['worst']['total_calls']),
            _format_decimal(payload['tipification']['worst']['retention_rate']),
        ])


def _build_acoes_sheet(workbook: Workbook, payload: dict, filters: dict = None):
    """Sheet 4: Resumo de ações."""
    ws = workbook.create_sheet('Acoes')

    headers = ['Acao', 'Total', 'Taxa sucesso (%)']
    ws.append(headers)
    _setup_worksheet_styles(ws)

    actions = payload['actions']
    rows = []

    if actions['most_used']:
        rows.append([
            f"{actions['most_used']['retention_action']} (Mais usada)",
            int(actions['most_used']['total_used']),
            _format_decimal(actions['most_used']['success_rate']),
        ])

    if actions['highest_success']:
        rows.append([
            f"{actions['highest_success']['retention_action']} (Maior sucesso)",
            int(actions['highest_success']['total_used']),
            _format_decimal(actions['highest_success']['success_rate']),
        ])

    if actions['lowest_success']:
        rows.append([
            f"{actions['lowest_success']['retention_action']} (Menor sucesso)",
            int(actions['lowest_success']['total_used']),
            _format_decimal(actions['lowest_success']['success_rate']),
        ])

    rows.append([
        'Sem acao',
        int(round(actions['no_action_pct'] * payload['kpis']['total_calls'] / 100, 0)),
        _format_decimal(actions['no_action_pct']),
    ])

    for row in rows:
        ws.append(row)


def _build_auditoria_sheet(workbook: Workbook, payload: dict):
    """Sheet 5: Chamadas priorizadas para auditoria."""
    ws = workbook.create_sheet('Auditoria')

    headers = [
        'ID Chamada',
        'Assistente',
        'Data',
        'Motivo corte',
        'Acao retencao',
        'Resultado final',
        'Razoes priorizacao',
    ]
    ws.append(headers)
    _setup_worksheet_styles(ws)

    for call in payload['audit_calls']:
        ws.append([
            call['call_id_external'],
            call['assistant_name'],
            call['occurred_on'].strftime('%Y-%m-%d') if call['occurred_on'] else '',
            call['churn_reason'],
            call['retention_action'],
            call['final_outcome'],
            ' | '.join(call['priority_reasons']),
        ])


def _build_insights_sheet(workbook: Workbook, payload: dict):
    """Sheet 6: Insights automáticos com ações sugeridas."""
    ws = workbook.create_sheet('Insights')

    headers = [
        'Titulo',
        'Resumo',
        'Interpretacao operacional',
        'Acoes sugeridas',
        'Recomendacao auditoria',
    ]
    ws.append(headers)
    _setup_worksheet_styles(ws)

    for insight in payload['insights']:
        suggested_actions = ' | '.join(insight.get('suggested_actions', [])) or '-'
        ws.append([
            insight.get('title', ''),
            insight.get('summary', ''),
            insight.get('operational_interpretation', ''),
            suggested_actions,
            insight.get('audit_recommendation', ''),
        ])

    # Aumenta altura das linhas para insights
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        ws.row_dimensions[row_idx].height = 50
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
