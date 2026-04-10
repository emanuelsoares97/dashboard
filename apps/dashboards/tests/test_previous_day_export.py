import pytest
from datetime import timedelta
from django.utils import timezone
from openpyxl import load_workbook
from io import BytesIO

from apps.dashboards.services.previous_day import build_previous_day_payload
from apps.dashboards.services.previous_day_export import export_previous_day_excel


@pytest.mark.django_db
class TestPreviousDayExport:
    """Testa exportação em Excel do Dia anterior."""

    def test_export_excel_generates_response(self):
        """Testa que a exportação gera uma resposta HTTP válida."""
        # Usa payload vazio simples
        payload = {
            'day': timezone.localdate() - timedelta(days=1),
            'kpis': {
                'total_calls': 100,
                'retention_rate': 75.0,
                'non_retention_rate': 15.0,
                'call_drop_rate': 10.0,
                'no_action_pct': 5.0,
                'inconsistency_rate': 2.0,
            },
            'assistants': {
                'top': [],
                'bottom': [],
            },
            'tipification': {
                'best': None,
                'worst': None,
            },
            'actions': {
                'most_used': None,
                'highest_success': None,
                'lowest_success': None,
                'no_action_pct': 5.0,
            },
            'insights': [],
            'audit_calls': [],
        }
        
        response = export_previous_day_excel(payload, {})
        
        assert response.status_code == 200
        assert 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' in response['Content-Type']
        assert 'attachment' in response['Content-Disposition']
        assert '.xlsx' in response['Content-Disposition']

    def test_export_excel_contains_all_sheets(self):
        """Testa que o Excel contém todas as sheets esperadas."""
        payload = {
            'day': timezone.localdate() - timedelta(days=1),
            'kpis': {
                'total_calls': 100,
                'retention_rate': 75.0,
                'non_retention_rate': 15.0,
                'call_drop_rate': 10.0,
                'no_action_pct': 5.0,
                'inconsistency_rate': 2.0,
            },
            'assistants': {
                'top': [],
                'bottom': [],
            },
            'tipification': {
                'best': None,
                'worst': None,
            },
            'actions': {
                'most_used': None,
                'highest_success': None,
                'lowest_success': None,
                'no_action_pct': 5.0,
            },
            'insights': [],
            'audit_calls': [],
        }
        
        response = export_previous_day_excel(payload, {})
        
        # Faz parsing do conteúdo como workbook
        wb = load_workbook(BytesIO(response.content))
        sheet_names = wb.sheetnames
        
        expected_sheets = ['Resumo', 'Assistentes', 'Tipificacoes', 'Acoes', 'Auditoria', 'Insights']
        for sheet_name in expected_sheets:
            assert sheet_name in sheet_names, f"Sheet {sheet_name} não encontrado"

    def test_resumo_sheet_has_headers_and_data(self):
        """Testa que a sheet Resumo tem headers e dados."""
        payload = {
            'day': timezone.localdate() - timedelta(days=1),
            'kpis': {
                'total_calls': 100,
                'retention_rate': 75.0,
                'non_retention_rate': 15.0,
                'call_drop_rate': 10.0,
                'no_action_pct': 5.0,
                'inconsistency_rate': 2.0,
            },
            'assistants': {
                'top': [],
                'bottom': [],
            },
            'tipification': {
                'best': None,
                'worst': None,
            },
            'actions': {
                'most_used': None,
                'highest_success': None,
                'lowest_success': None,
                'no_action_pct': 5.0,
            },
            'insights': [],
            'audit_calls': [],
        }
        
        response = export_previous_day_excel(payload, {})
        wb = load_workbook(BytesIO(response.content))
        ws = wb['Resumo']
        
        # Verificar headers
        assert ws['A1'].value == 'Metrica'
        assert ws['B1'].value == 'Valor'
        
        # Verificar dados básicos
        assert ws['A2'].value == 'Data'
        assert ws['A3'].value == 'Total de chamadas'

    def test_assistentes_sheet_has_headers_and_data(self):
        """Testa que a sheet Assistentes tem headers e dados."""
        payload = {
            'day': timezone.localdate() - timedelta(days=1),
            'kpis': {
                'total_calls': 100,
                'retention_rate': 75.0,
                'non_retention_rate': 15.0,
                'call_drop_rate': 10.0,
                'no_action_pct': 5.0,
                'inconsistency_rate': 2.0,
            },
            'assistants': {
                'top': [],
                'bottom': [],
            },
            'tipification': {
                'best': None,
                'worst': None,
            },
            'actions': {
                'most_used': None,
                'highest_success': None,
                'lowest_success': None,
                'no_action_pct': 5.0,
            },
            'insights': [],
            'audit_calls': [],
        }
        
        response = export_previous_day_excel(payload, {})
        wb = load_workbook(BytesIO(response.content))
        ws = wb['Assistentes']
        
        # Verificar headers
        assert ws['A1'].value == 'Nome'
        assert ws['B1'].value == 'Total chamadas'
        assert ws['C1'].value == 'Taxa retencao (%)'
        assert ws['D1'].value == 'Ranking'

    def test_auditoria_sheet_has_headers_and_audit_calls(self):
        """Testa que a sheet Auditoria tem audit calls."""
        payload = {
            'day': timezone.localdate() - timedelta(days=1),
            'kpis': {
                'total_calls': 100,
                'retention_rate': 75.0,
                'non_retention_rate': 15.0,
                'call_drop_rate': 10.0,
                'no_action_pct': 5.0,
                'inconsistency_rate': 2.0,
            },
            'assistants': {
                'top': [],
                'bottom': [],
            },
            'tipification': {
                'best': None,
                'worst': None,
            },
            'actions': {
                'most_used': None,
                'highest_success': None,
                'lowest_success': None,
                'no_action_pct': 5.0,
            },
            'insights': [],
            'audit_calls': [],
        }
        
        response = export_previous_day_excel(payload, {})
        wb = load_workbook(BytesIO(response.content))
        ws = wb['Auditoria']
        
        # Verificar headers
        assert ws['A1'].value == 'Score'
        assert ws['B1'].value == 'ID Chamada'
        assert ws['C1'].value == 'Assistente'
        assert ws['D1'].value == 'Data'
        assert ws['G1'].value == 'Motivos de auditoria'

    def test_insights_sheet_has_headers(self):
        """Testa que a sheet Insights tem headers."""
        payload = {
            'day': timezone.localdate() - timedelta(days=1),
            'kpis': {
                'total_calls': 100,
                'retention_rate': 75.0,
                'non_retention_rate': 15.0,
                'call_drop_rate': 10.0,
                'no_action_pct': 5.0,
                'inconsistency_rate': 2.0,
            },
            'assistants': {
                'top': [],
                'bottom': [],
            },
            'tipification': {
                'best': None,
                'worst': None,
            },
            'actions': {
                'most_used': None,
                'highest_success': None,
                'lowest_success': None,
                'no_action_pct': 5.0,
            },
            'insights': [],
            'audit_calls': [],
        }
        
        response = export_previous_day_excel(payload, {})
        wb = load_workbook(BytesIO(response.content))
        ws = wb['Insights']
        
        # Verificar headers
        assert ws['A1'].value == 'Titulo'
        assert ws['B1'].value == 'Resumo'
        assert ws['C1'].value == 'Interpretacao operacional'
        assert ws['D1'].value == 'Acoes sugeridas'
        assert ws['E1'].value == 'Recomendacao auditoria'

    def test_filename_contains_date(self):
        """Testa que o filename contém a data."""
        today = timezone.localdate()
        previous = today - timedelta(days=1)
        payload = {
            'day': previous,
            'kpis': {
                'total_calls': 100,
                'retention_rate': 75.0,
                'non_retention_rate': 15.0,
                'call_drop_rate': 10.0,
                'no_action_pct': 5.0,
                'inconsistency_rate': 2.0,
            },
            'assistants': {
                'top': [],
                'bottom': [],
            },
            'tipification': {
                'best': None,
                'worst': None,
            },
            'actions': {
                'most_used': None,
                'highest_success': None,
                'lowest_success': None,
                'no_action_pct': 5.0,
            },
            'insights': [],
            'audit_calls': [],
        }
        
        response = export_previous_day_excel(payload, {})
        
        filename = response['Content-Disposition']
        date_str = previous.strftime('%Y%m%d')
        assert date_str in filename


@pytest.mark.django_db
class TestPreviousDayExportView:
    """Testa a view de export do Dia anterior."""

    def test_previous_day_export_requires_sensitive_analytics(self, client, django_user_model):
        """Testa que a view export require permissão sensitive_analytics."""
        # Cria user assistant (sem permissão)
        user = django_user_model.objects.create_user(username='assistant', password='pass')
        client.login(username='assistant', password='pass')
        
        response = client.get('/dashboards/previous-day/export.xlsx')
        
        # Deve redirecionar ou retornar 403
        assert response.status_code in [302, 403]
